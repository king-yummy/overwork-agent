# -*- coding: utf-8 -*-
"""Stage 6 — 배포 파일(초과근무_팀별) 템플릿 채우기.

기존 템플릿의 데이터 시트에만 값을 써넣고(차트·수식·서식 보존),
가공/안내자료/그래프/전년비교는 템플릿 수식이 자동 계산하도록 둔다.

데이터 시트 매핑:
  · 1)인별_복사_가공(a~c)  : 4행~, D=사번 E=성명 G=근무코드 H=근무부서, L~W=1~12월
  · 2)팀전체_복사          : 3행~, A~H=담당Ⅲ/Ⅱ/Ⅰ+근무부서코드/명, J~U=월평균, V=연합(총초근), W=연평균
  · 수신인                 : 2행~, A=조직코드 B=조직명 C=사번 D=수신인 E=이메일
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from xlsx_surgery import fill_template

SHEET_INBYEOL = "1)인별_복사_가공(a~c)"
SHEET_TEAM = "2)팀전체_복사"
SHEET_SUSIN = "수신인"


def _L(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _resolve(template_path, prefix):
    """템플릿의 실제 시트명을 prefix로 찾는다 ('1)인별_복사_가공(a~c)' 같은 변형 대비)."""
    wb = load_workbook(template_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    for n in names:
        if n.replace(" ", "").startswith(prefix.replace(" ", "")):
            return n
    raise KeyError(f"템플릿에 '{prefix}…' 시트가 없습니다. 실제 시트: {names}")


def _maxrow(path, sheet):
    wb = load_workbook(path, read_only=True)
    mr = wb[sheet].max_row
    wb.close()
    return mr


def _mnum(m):           # '5월' -> 5
    return int(str(m).replace("월", "").strip())


def _read_prior(path, sheet, key_col, data_row, base_col, upto_month):
    """기존 배포 파일에서 이전 월(1..upto_month) 값을 key→{월:값}으로 읽는다."""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    out, oldV = {}, {}
    for r in range(data_row, ws.max_row + 1):
        key = ws.cell(row=r, column=key_col).value
        if key in (None, ""):
            continue
        key = str(key).strip()
        out[key] = {m: (ws.cell(row=r, column=base_col + m).value or 0) for m in range(1, upto_month)}
    wb.close()
    return out


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def build_team_file(template_path, out_path, 인별, 평균, 총초근, 수신인, 근무월):
    cur = _mnum(근무월)                       # 이번 달 (예: 5)
    sh_i = _resolve(template_path, "1)인별_복사_가공")
    sh_t = _resolve(template_path, "2)팀전체_복사")
    sh_s = _resolve(template_path, "수신인")

    # 이전 월(1..cur-1)을 기존 파일에서 캐리
    prior_i = _read_prior(template_path, sh_i, 4, 4, 11, cur)      # 인별: 사번(D=4), 1월=L(12)
    wb = load_workbook(template_path, data_only=True)              # 팀전체: 근무부서코드(G=7), 1월=J(10), V=22
    wst = wb[sh_t]
    prior_t, oldV = {}, {}
    for r in range(3, wst.max_row + 1):
        code = wst.cell(row=r, column=7).value
        if code in (None, ""):
            continue
        code = str(code).strip()
        prior_t[code] = {m: _num(wst.cell(row=r, column=9 + m).value) for m in range(1, cur)}
        oldV[code] = _num(wst.cell(row=r, column=22).value)
    wb.close()

    # ── 1)인별_복사_가공 : 현재 재직추출 행 + (이전월 캐리 / 이번달 계산) ──
    DATA_COLS_I = {"D", "E", "G", "H", "I", "K"} | {_L(11 + k) for k in range(1, 13)}
    # 템플릿 4행에서 '행별 수식 컬럼' 감지 (A,B,C,X.. 등 — 내가 값으로 쓰는 칸 제외)
    wbf = load_workbook(template_path)
    wsf = wbf[sh_i]
    formula_cols = {}
    for c in range(1, wsf.max_column + 1):
        v = wsf.cell(row=4, column=c).value
        if isinstance(v, str) and v.startswith("=") and _L(c) not in DATA_COLS_I:
            formula_cols[_L(c)] = v
    wbf.close()

    cells_i = {}
    r = 4
    for _, row in 인별.iterrows():
        sab = str(row["사번"])
        cells_i[("D", r)] = (sab, True)
        cells_i[("E", r)] = (row["성명"], True)
        cells_i[("G", r)] = (str(row["근무코드"]), True)
        cells_i[("H", r)] = (row["근무부서"], True)
        for m in range(1, cur):                                   # 이전 월: 기존 파일에서
            cells_i[(_L(11 + m), r)] = (float(prior_i.get(sab, {}).get(m, 0) or 0), False)
        cells_i[(_L(11 + cur), r)] = (float(row.get(근무월, 0) or 0), False)  # 이번 달: 엔진
        for col, f0 in formula_cols.items():                      # 행별 수식 복제
            cells_i[(col, r)] = (Translator(f0, origin=f"{col}4").translate_formula(f"{col}{r}"), "f")
        r += 1
    last_i = r - 1
    # 2행 요약: 총인원(D2)·월별 총합(L2~W2) — 그래프/가공의 월 게이트($..$4>0)가 이 값을 본다
    cells_i[("D", 2)] = (f"=COUNTA(D4:D{last_i})", "f")
    for k in range(1, 13):
        col = _L(11 + k)
        cells_i[(col, 2)] = (f"=SUM({col}4:{col}{last_i})", "f")
    for rr in range(last_i + 1, _maxrow(template_path, sh_i) + 1):   # 잉여행: 데이터+수식 모두 비움
        for col in list(DATA_COLS_I) + list(formula_cols.keys()):
            cells_i[(col, rr)] = ("", False)

    # ── 2)팀전체_복사 : 이전월 캐리 + 이번달 + 연합/연평균 재계산(1~cur) ──
    tot = 총초근.set_index("근무부서코드")
    cells_t = {}
    r = 3
    for _, row in 평균.iterrows():
        code = str(row["근무부서코드"])
        cells_t[("A", r)] = (str(row["담당3코드"]), True)
        cells_t[("B", r)] = (row["담당3"], True)
        cells_t[("C", r)] = (str(row["담당2코드"]), True)
        cells_t[("D", r)] = (row["담당2"], True)
        cells_t[("E", r)] = (str(row["담당1코드"]), True)
        cells_t[("F", r)] = (row["담당1"], True)
        cells_t[("G", r)] = (code, True)
        cells_t[("H", r)] = (row["근무부서명"], True)
        월평균 = []
        for m in range(1, cur):                                   # 이전 월 평균: 캐리
            v = float(prior_t.get(code, {}).get(m, 0) or 0)
            cells_t[(_L(9 + m), r)] = (v, False); 월평균.append(v)
        cur_avg = float(row.get(근무월, 0) or 0)                   # 이번 달 평균: 엔진
        cells_t[(_L(9 + cur), r)] = (cur_avg, False); 월평균.append(cur_avg)
        cur_tot = float(tot.loc[code, 근무월]) if code in tot.index else 0.0
        cells_t[("V", r)] = (oldV.get(code, 0.0) + cur_tot, False)       # 연합(총초근) 1~cur
        cells_t[("W", r)] = (sum(월평균) / len(월평균) if 월평균 else 0.0, False)  # 연평균
        r += 1
    last_t = r - 1
    for rr in range(last_t + 1, _maxrow(template_path, sh_t) + 1):
        for col in list("ABCDEFGH") + [_L(9 + k) for k in range(1, 13)] + ["V", "W"]:
            cells_t[(col, rr)] = ("", False)

    # ── 수신인 ──
    cells_s = {}
    r = 2
    for _, row in 수신인.iterrows():
        cells_s[("A", r)] = (str(row["조직코드"]), True)
        cells_s[("B", r)] = (row["조직명"], True)
        cells_s[("C", r)] = (str(row["사번"]), True)
        cells_s[("D", r)] = (row["수신인"], True)
        cells_s[("E", r)] = (row["이메일"], True)
        r += 1
    for rr in range(r, _maxrow(template_path, sh_s) + 1):
        for col in "ABCDE":
            cells_s[(col, rr)] = ("", False)

    fill_template(template_path, out_path, {
        sh_i: cells_i, sh_t: cells_t, sh_s: cells_s,
    })
    return out_path


def _inbyeol_cells(template_path, sh_i, 인별, prior_i, cur, 근무월):
    """1)인별_복사_가공 셀 dict (팀별·담당별 공용)."""
    DATA = {"D", "E", "G", "H", "I", "K"} | {_L(11 + k) for k in range(1, 13)}
    wbf = load_workbook(template_path); wsf = wbf[sh_i]
    fcols = {}
    for c in range(1, wsf.max_column + 1):
        v = wsf.cell(row=4, column=c).value
        if isinstance(v, str) and v.startswith("=") and _L(c) not in DATA:
            fcols[_L(c)] = v
    wbf.close()
    cells = {}; r = 4
    for _, row in 인별.iterrows():
        sab = str(row["사번"])
        cells[("D", r)] = (sab, True); cells[("E", r)] = (row["성명"], True)
        cells[("G", r)] = (str(row["근무코드"]), True); cells[("H", r)] = (row["근무부서"], True)
        for m in range(1, cur):
            cells[(_L(11 + m), r)] = (float(prior_i.get(sab, {}).get(m, 0) or 0), False)
        cells[(_L(11 + cur), r)] = (float(row.get(근무월, 0) or 0), False)
        for col, f0 in fcols.items():
            cells[(col, r)] = (Translator(f0, origin=f"{col}4").translate_formula(f"{col}{r}"), "f")
        r += 1
    last = r - 1
    cells[("D", 2)] = (f"=COUNTA(D4:D{last})", "f")
    for k in range(1, 13):
        col = _L(11 + k); cells[(col, 2)] = (f"=SUM({col}4:{col}{last})", "f")
    for rr in range(last + 1, _maxrow(template_path, sh_i) + 1):
        for col in list(DATA) + list(fcols.keys()):
            cells[(col, rr)] = ("", False)
    return cells


def build_dept_file(template_path, out_path, 인별, 팀평균, 팀총초근,
                    담당평균, 담당총초근, 수신인, 근무월, org_table):
    cur = _mnum(근무월)
    sh_i = _resolve(template_path, "1)인별_복사_가공")
    sh_t = _resolve(template_path, "2)팀전체_복사")
    sh_d = _resolve(template_path, "3)담당전체_복사")
    sh_s = _resolve(template_path, "수신인")

    prior_i = _read_prior(template_path, sh_i, 4, 4, 11, cur)
    wb = load_workbook(template_path, data_only=True)
    wt = wb[sh_t]; prior_t = {}; oldVt = {}
    for r in range(3, wt.max_row + 1):              # 2)팀전체(담당별): 근무부서코드 J=10, 1월 M=13, 연합 Y=25
        k = wt.cell(row=r, column=10).value
        if k in (None, ""): continue
        k = str(k).strip()
        prior_t[k] = {m: _num(wt.cell(row=r, column=12 + m).value) for m in range(1, cur)}
        oldVt[k] = _num(wt.cell(row=r, column=25).value)
    wd = wb[sh_d]; prior_d = {}; oldVd = {}
    for r in range(3, wd.max_row + 1):              # 3)담당전체: 담당Ⅰ코드 E=5, 1월 G=7, 연합 S=19
        k = wd.cell(row=r, column=5).value
        if k in (None, ""): continue
        k = str(k).strip()
        prior_d[k] = {m: _num(wd.cell(row=r, column=6 + m).value) for m in range(1, cur)}
        oldVd[k] = _num(wd.cell(row=r, column=19).value)
    wb.close()

    cells_i = _inbyeol_cells(template_path, sh_i, 인별, prior_i, cur, 근무월)

    # 2)팀전체_복사 (담당별: A=I&B, B=COUNTIFS, C=순번, D~K 계층, M~X 월평균, Y연합, Z연평균)
    tot = 팀총초근.set_index("근무부서코드")
    cells_t = {}; seq = 1; r = 3
    for _, row in 팀평균.iterrows():
        code = str(row["근무부서코드"])
        cells_t[("A", r)] = (f"=I{r}&B{r}", "f")
        cells_t[("B", r)] = (f'=COUNTIFS($H$3:$H$2413,H{r},$C$3:$C$2413,"<"&C{r})+1', "f")
        cells_t[("C", r)] = (seq, False)
        cells_t[("D", r)] = (str(row["담당3코드"]), True); cells_t[("E", r)] = (row["담당3"], True)
        cells_t[("F", r)] = (str(row["담당2코드"]), True); cells_t[("G", r)] = (row["담당2"], True)
        cells_t[("H", r)] = (str(row["담당1코드"]), True); cells_t[("I", r)] = (row["담당1"], True)
        cells_t[("J", r)] = (code, True); cells_t[("K", r)] = (row["근무부서명"], True)
        월평균 = []
        for m in range(1, cur):
            v = float(prior_t.get(code, {}).get(m, 0) or 0); cells_t[(_L(12 + m), r)] = (v, False); 월평균.append(v)
        ca = float(row.get(근무월, 0) or 0); cells_t[(_L(12 + cur), r)] = (ca, False); 월평균.append(ca)
        ct = float(tot.loc[code, 근무월]) if code in tot.index else 0.0
        cells_t[("Y", r)] = (oldVt.get(code, 0.0) + ct, False)
        cells_t[("Z", r)] = (sum(월평균) / len(월평균) if 월평균 else 0.0, False)
        seq += 1; r += 1
    last_t = r - 1
    for rr in range(last_t + 1, _maxrow(template_path, sh_t) + 1):
        for col in list("ABCDEFGHIJK") + [_L(12 + k) for k in range(1, 13)] + ["Y", "Z"]:
            cells_t[(col, rr)] = ("", False)

    # 3)담당전체_복사 (담당Ⅰ 롤업: A~F 계층, G~R 월평균, S연합, T연평균)
    htab = {}
    for _, o in org_table.iterrows():
        k = str(o["담당1코드"])
        if k not in htab:
            htab[k] = (str(o["담당2코드"]), o["담당2명"], str(o["담당3코드"]), o["담당3명"])
    dtot = 담당총초근.set_index("담당1코드")
    cells_d = {}; r = 3
    for _, row in 담당평균.iterrows():
        code = str(row["담당1코드"])
        d2c, d2n, d3c, d3n = htab.get(code, ("", "", "", ""))
        cells_d[("A", r)] = (d3c, True); cells_d[("B", r)] = (d3n, True)
        cells_d[("C", r)] = (d2c, True); cells_d[("D", r)] = (d2n, True)
        cells_d[("E", r)] = (code, True); cells_d[("F", r)] = (row["담당1"], True)
        월평균 = []
        for m in range(1, cur):
            v = float(prior_d.get(code, {}).get(m, 0) or 0); cells_d[(_L(6 + m), r)] = (v, False); 월평균.append(v)
        ca = float(row.get(근무월, 0) or 0); cells_d[(_L(6 + cur), r)] = (ca, False); 월평균.append(ca)
        ct = float(dtot.loc[code, 근무월]) if code in dtot.index else 0.0
        cells_d[("S", r)] = (oldVd.get(code, 0.0) + ct, False)
        cells_d[("T", r)] = (sum(월평균) / len(월평균) if 월평균 else 0.0, False)
        r += 1
    last_d = r - 1
    for rr in range(last_d + 1, _maxrow(template_path, sh_d) + 1):
        for col in list("ABCDEF") + [_L(6 + k) for k in range(1, 13)] + ["S", "T"]:
            cells_d[(col, rr)] = ("", False)

    # 수신인
    cells_s = {}; r = 2
    for _, row in 수신인.iterrows():
        cells_s[("A", r)] = (str(row["조직코드"]), True); cells_s[("B", r)] = (row["조직명"], True)
        cells_s[("C", r)] = (str(row["사번"]), True); cells_s[("D", r)] = (row["수신인"], True)
        cells_s[("E", r)] = (row["이메일"], True); r += 1
    for rr in range(r, _maxrow(template_path, sh_s) + 1):
        for col in "ABCDE":
            cells_s[(col, rr)] = ("", False)

    # 1행 인평균(= 인별 월총합 / 인별 인원) — 그래프 S~Z 도표의 월 게이트($..$4>0)가 이 값을 본다
    qi = f"'{sh_i}'"
    for k in range(1, 13):
        cells_t[(_L(12 + k), 1)] = (f"={qi}!{_L(11 + k)}2/{qi}!$D$2", "f")  # 2)팀전체 M1~X1
        cells_d[(_L(6 + k), 1)] = (f"={qi}!{_L(11 + k)}2/{qi}!$D$2", "f")   # 3)담당전체 G1~R1

    fill_template(template_path, out_path,
                  {sh_i: cells_i, sh_t: cells_t, sh_d: cells_d, sh_s: cells_s})
    return out_path